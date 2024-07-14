import pandas as pd
from openpyxl import load_workbook


class FAnalyzer:
    def __init__(self, filename):
        self.origin_df = None
        self.point_map = None
        self.filename = filename

    def analyze(self):
        # 设置显示更多的行和列
        pd.set_option('display.max_rows', 100)  # 显示最多 100 行
        pd.set_option('display.max_columns', 20)  # 显示最多 20 列

        df_map = pd.read_excel(self.filename, sheet_name='Sheet2', header=[0])
        self.point_map = pd.Series(df_map.iloc[:, 1].values, index=df_map.iloc[:, 0]).to_dict()
        print("[Step1]转换后的字典长度:", len(self.point_map))

        self.origin_df = pd.read_excel(self.filename, sheet_name='Sheet1', header=[0])
        print("[Step2]输入的列数量为:", len(self.origin_df.columns))

        self.origin_df['得分'] = self.origin_df.apply(self.process_row, axis=1)
        print("[Step3]获取到得分数量:", len(self.origin_df['得分']))

        # 加载现有的 Excel 文件
        wb = load_workbook(self.filename)
        ws = wb['Sheet1']
        # 获取最后一列的列号
        last_col_letter = ws.cell(row=1, column=ws.max_column).column_letter
        # 将 '结果' 列写回到 Excel 文件的最后一列，并保留下拉列表
        for index, value in enumerate(self.origin_df['得分'], start=2):
            ws[f'{last_col_letter}{index}'].value = value
        wb.save(self.filename)

        print("[Step4]处理完成，并将结果写回")

    def process_row(self, row):
        total = 0
        for col in self.origin_df.columns:
            if col == '代码' or col == '名称' or col == '得分':
                continue

            if row[col] in self.point_map:
                row[col] = self.point_map[row[col]]
                total += row[col]
            else:
                print('Missed Mapping: ', row[col], ' col= ', col)
        return total



